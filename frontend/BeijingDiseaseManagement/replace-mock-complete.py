#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import re
import json
from pathlib import Path

# 读取Excel数据
from openpyxl import load_workbook

excel_file = Path('src/views/patient/data1000.xlsx')
wb = load_workbook(excel_file)
ws = wb.active

# 获取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

# 获取前100条数据
data = []
for row_idx in range(2, min(102, ws.max_row + 1)):
    row_data = {}
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        row_data[header] = cell.value
    data.append(row_data)

# 读取原始文件
vue_file = Path('src/views/patient/his-table.vue')
content = vue_file.read_text(encoding='utf-8')

# 生成新的mock数据
mock_data_json = json.dumps(data, ensure_ascii=False, indent=4)
new_mock_code = f"  // Mock 数据 - 从Excel文件导入的100条患者记录\n  const mockData = {mock_data_json};"

# 使用正则表达式替换mock数据部分
# 匹配从 "// Mock 数据" 开始到 "];" 结束的部分
pattern = r'  // Mock 数据.*?\n  const mockData = \[[\s\S]*?\n  \];'
content_new = re.sub(pattern, new_mock_code, content, count=1)

# 写回文件
vue_file.write_text(content_new, encoding='utf-8')

print('Successfully replaced mock data with 100 records from Excel')
print(f'Total records: {len(data)}')



